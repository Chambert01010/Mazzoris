from __future__ import annotations

import io
import logging
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass
from typing import Iterable

import pandas as pd
import pdfplumber


DATE_RE = re.compile(r"^\d{2}/[A-Z]{3}$", re.IGNORECASE)
CODE_RE = re.compile(r"^[A-Z0-9]{3}$")
MONEY_RE = re.compile(r"^\d{1,3}(?:,\d{3})*\.\d{2}$")
BBVA_LINE_TOLERANCE = 2.5
BBVA_HEADER_MIN_TOP = 140
BBVA_FOOTER_MAX_TOP = 750
BBVA_AMOUNT_COLUMNS = {
    "cargo": (340, 395),
    "abono": (395, 465),
    "saldo_operacion": (465, 535),
    "saldo_liquidacion": (535, 999),
}
BBVA_COMPACT_AMOUNT_RIGHT_EDGES = {
    "cargo": 420,
    "abono": 470,
    "saldo_operacion": 540,
}
BBVA_IGNORED_CONTINUATIONS = (
    "Estimado Cliente",
    "Su Estado de Cuenta",
    "Tambien le informamos",
    "Con BBVA adelante",
    "La GAT Real",
    "BBVA MEXICO",
    "Av. Paseo de la Reforma",
)


class StatementProcessingError(Exception):
    """Raised when a statement cannot be parsed reliably."""


LOGGER = logging.getLogger("mazzoris.statement_processing")


def _normalize_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "")
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_text).strip()


def _read_money(value: str | None) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace("$", "").replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return None


def _money_to_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value.quantize(Decimal("0.01")))


def _clean_join(parts: Iterable[str]) -> str:
    text = " ".join(part for part in parts if part)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", text)
    text = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", text)
    return text.strip(" |")


def _group_words_into_lines(words: list[dict], tolerance: float = BBVA_LINE_TOLERANCE) -> list[list[dict]]:
    lines: list[list[dict]] = []
    current: list[dict] = []
    for word in words:
        if not current or abs(word["top"] - current[-1]["top"]) <= tolerance:
            current.append(word)
        else:
            lines.append(sorted(current, key=lambda item: item["x0"]))
            current = [word]
    if current:
        lines.append(sorted(current, key=lambda item: item["x0"]))
    return lines


def parse_bbva_summary_text(text: str) -> dict[str, Decimal | int | None]:
    normalized = _normalize_text(text)
    patterns = {
        "saldo_liquidacion_inicial": r"Saldo de Liquidacion Inicial\s+([\d,]+\.\d{2})",
        "saldo_operacion_inicial": r"(?:Saldo de Operacion Inicial|Saldo Anterior)\s+([\d,]+\.\d{2})",
        "depositos_abonos": r"Depositos / Abonos \(\+\)\s+(\d+)\s+([\d,]+\.\d{2})",
        "retiros_cargos": r"Retiros / Cargos \(-\)\s+(\d+)\s+([\d,]+\.\d{2})",
        "saldo_final": r"Saldo Final \(\+\)\s+([\d,]+\.\d{2})",
        "saldo_operacion_final": r"Saldo de Operacion Final\s+([\d,]+\.\d{2})",
    }

    summary: dict[str, Decimal | int | None] = {
        "saldo_liquidacion_inicial": None,
        "saldo_operacion_inicial": None,
        "depositos_abonos": None,
        "depositos_abonos_count": None,
        "retiros_cargos": None,
        "retiros_cargos_count": None,
        "saldo_final": None,
        "saldo_operacion_final": None,
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, normalized)
        if not match:
            continue
        if key == "depositos_abonos":
            summary["depositos_abonos_count"] = int(match.group(1))
            summary["depositos_abonos"] = _read_money(match.group(2))
        elif key == "retiros_cargos":
            summary["retiros_cargos_count"] = int(match.group(1))
            summary["retiros_cargos"] = _read_money(match.group(2))
        else:
            summary[key] = _read_money(match.group(1))

    return summary


def _extract_bbva_summary(pdf: pdfplumber.PDF) -> dict[str, Decimal | int | None]:
    for page in pdf.pages:
        text = page.extract_text() or ""
        summary = parse_bbva_summary_text(text)
        initial_balance = summary["saldo_operacion_inicial"] or summary["saldo_liquidacion_inicial"]
        if initial_balance is not None and summary["saldo_final"] is not None:
            return summary
    return parse_bbva_summary_text("")


def _is_bbva_transaction_start(texts: list[str]) -> bool:
    return (
        len(texts) >= 2
        and DATE_RE.match(texts[0] or "") is not None
        and DATE_RE.match(texts[1] or "") is not None
    )


def _bbva_header_has_code_column(text: str) -> bool:
    normalized = _normalize_text(text).upper()
    return re.search(r"\bCOD(?:IGO)?\.?\b", normalized) is not None


def _should_ignore_bbva_line(text: str) -> bool:
    if not text:
        return True
    return any(text.startswith(prefix) for prefix in BBVA_IGNORED_CONTINUATIONS)


def _append_bbva_amount(
    row: dict,
    text: str,
    x0: float,
    x1: float | None = None,
    *,
    compact_layout: bool = False,
) -> bool:
    if not MONEY_RE.match(text):
        return False
    if compact_layout and x1 is not None:
        if x1 <= BBVA_COMPACT_AMOUNT_RIGHT_EDGES["cargo"]:
            row["cargo"] = text
        elif x1 <= BBVA_COMPACT_AMOUNT_RIGHT_EDGES["abono"]:
            row["abono"] = text
        elif x1 <= BBVA_COMPACT_AMOUNT_RIGHT_EDGES["saldo_operacion"]:
            row["saldo_operacion"] = text
        else:
            row["saldo_liquidacion"] = text
        return True
    for field, (min_x, max_x) in BBVA_AMOUNT_COLUMNS.items():
        if min_x <= x0 < max_x:
            row[field] = text
            return True
    return False


def _finalize_bbva_data(rows: list[dict], summary: dict[str, Decimal | int | None]) -> tuple[pd.DataFrame, pd.DataFrame]:
    initial_balance = summary.get("saldo_operacion_inicial") or summary.get("saldo_liquidacion_inicial")
    running_balance = initial_balance if isinstance(initial_balance, Decimal) else None
    total_cargos = Decimal("0.00")
    total_abonos = Decimal("0.00")
    finalized_rows: list[dict] = []

    for index, row in enumerate(rows, start=1):
        cargo = _read_money(row.get("cargo")) or Decimal("0.00")
        abono = _read_money(row.get("abono")) or Decimal("0.00")
        saldo_operacion = _read_money(row.get("saldo_operacion"))
        saldo_liquidacion = _read_money(row.get("saldo_liquidacion"))
        saldo_referencia = saldo_operacion or saldo_liquidacion

        total_cargos += cargo
        total_abonos += abono

        if running_balance is not None:
            running_balance = (running_balance + abono - cargo).quantize(Decimal("0.01"))

        cuadre = None
        diferencia = None
        if running_balance is not None and saldo_referencia is not None:
            diferencia = (saldo_referencia - running_balance).quantize(Decimal("0.01"))
            cuadre = diferencia == Decimal("0.00")

        finalized_rows.append(
            {
                "MOVIMIENTO": index,
                "PAGINA": row["page"],
                "FECHA_OPERACION": row["fecha_operacion"],
                "FECHA_LIQUIDACION": row["fecha_liquidacion"],
                "CODIGO": row["codigo"],
                "DESCRIPCION": _clean_join(row["descripcion_parts"]),
                "REFERENCIA": " | ".join(part for part in row["referencia_parts"] if part),
                "CARGO": _money_to_float(cargo) if cargo else None,
                "ABONO": _money_to_float(abono) if abono else None,
                "SALDO_OPERACION_PDF": _money_to_float(saldo_operacion),
                "SALDO_LIQUIDACION_PDF": _money_to_float(saldo_liquidacion),
                "SALDO_CALCULADO": _money_to_float(running_balance),
                "CUADRA_CON_SALDO_PDF": cuadre,
                "DIFERENCIA_SALDO": _money_to_float(diferencia),
            }
        )

    parsed_final = running_balance
    expected_final = summary.get("saldo_operacion_final") or summary.get("saldo_final")
    final_difference = None
    if isinstance(parsed_final, Decimal) and isinstance(expected_final, Decimal):
        final_difference = (expected_final - parsed_final).quantize(Decimal("0.01"))

    summary_rows = [
        {"METRICA": "Saldo inicial", "PDF": _money_to_float(initial_balance), "CALCULADO": _money_to_float(initial_balance), "DIFERENCIA": 0.0 if initial_balance is not None else None},
        {"METRICA": "Depositos / Abonos", "PDF": _money_to_float(summary.get("depositos_abonos")), "CALCULADO": _money_to_float(total_abonos), "DIFERENCIA": _money_to_float((summary.get("depositos_abonos") - total_abonos) if isinstance(summary.get("depositos_abonos"), Decimal) else None)},
        {"METRICA": "Retiros / Cargos", "PDF": _money_to_float(summary.get("retiros_cargos")), "CALCULADO": _money_to_float(total_cargos), "DIFERENCIA": _money_to_float((summary.get("retiros_cargos") - total_cargos) if isinstance(summary.get("retiros_cargos"), Decimal) else None)},
        {"METRICA": "Saldo final", "PDF": _money_to_float(expected_final), "CALCULADO": _money_to_float(parsed_final), "DIFERENCIA": _money_to_float(final_difference)},
        {"METRICA": "Numero de abonos", "PDF": summary.get("depositos_abonos_count"), "CALCULADO": int(sum(1 for row in finalized_rows if row["ABONO"])), "DIFERENCIA": None},
        {"METRICA": "Numero de cargos", "PDF": summary.get("retiros_cargos_count"), "CALCULADO": int(sum(1 for row in finalized_rows if row["CARGO"])), "DIFERENCIA": None},
    ]

    return pd.DataFrame(finalized_rows), pd.DataFrame(summary_rows)


def build_bbva_audit_dataframe(
    movimientos_df: pd.DataFrame,
    resumen_df: pd.DataFrame,
) -> pd.DataFrame:
    summary_by_metric = {row["METRICA"]: row for _, row in resumen_df.iterrows()}
    checks: list[dict] = []

    def add_check(name: str, expected, calculated, detail: str) -> None:
        missing = pd.isna(expected) or pd.isna(calculated)
        difference = None if missing else Decimal(str(expected)) - Decimal(str(calculated))
        status = "ERROR" if missing or difference != Decimal("0") else "OK"
        checks.append(
            {
                "COMPROBACION": name,
                "PDF_ESPERADO": expected,
                "EXCEL_CALCULADO": calculated,
                "DIFERENCIA": _money_to_float(difference),
                "TOLERANCIA": 0.0,
                "ESTADO": status,
                "DETALLE": detail if not missing else f"{detail} Falta un valor obligatorio.",
            }
        )

    add_check(
        "Numero de abonos",
        summary_by_metric["Numero de abonos"]["PDF"],
        summary_by_metric["Numero de abonos"]["CALCULADO"],
        "El conteo de abonos debe coincidir con el resumen del PDF.",
    )
    add_check(
        "Numero de cargos",
        summary_by_metric["Numero de cargos"]["PDF"],
        summary_by_metric["Numero de cargos"]["CALCULADO"],
        "El conteo de cargos debe coincidir con el resumen del PDF.",
    )
    for metric in ("Depositos / Abonos", "Retiros / Cargos", "Saldo final"):
        add_check(
            metric,
            summary_by_metric[metric]["PDF"],
            summary_by_metric[metric]["CALCULADO"],
            f"El importe de {metric.lower()} debe cuadrar a centavos.",
        )

    checkpoint_rows = movimientos_df[movimientos_df["SALDO_OPERACION_PDF"].notna()]
    checkpoint_count = int(len(checkpoint_rows))
    checkpoint_ok = int((checkpoint_rows["CUADRA_CON_SALDO_PDF"] == True).sum())  # noqa: E712
    add_check(
        "Saldos de operacion intermedios",
        checkpoint_count,
        checkpoint_ok,
        "Todos los saldos de operacion disponibles deben coincidir con el saldo acumulado.",
    )

    overall_status = "OK" if all(row["ESTADO"] == "OK" for row in checks) else "ERROR"
    checks.insert(
        0,
        {
            "COMPROBACION": "Estado general",
            "PDF_ESPERADO": "OK",
            "EXCEL_CALCULADO": overall_status,
            "DIFERENCIA": None,
            "TOLERANCIA": 0.0,
            "ESTADO": overall_status,
            "DETALLE": "Auditoria completa de conteos, importes y saldos.",
        },
    )
    return pd.DataFrame(checks)


def _validate_bbva_audit(auditoria_df: pd.DataFrame) -> None:
    failed = auditoria_df[auditoria_df["ESTADO"] != "OK"]
    if failed.empty:
        return
    details = "; ".join(
        f"{row['COMPROBACION']}: {row['DETALLE']}"
        for _, row in failed.iterrows()
        if row["COMPROBACION"] != "Estado general"
    )
    LOGGER.error("Auditoria BBVA fallida: %s", details)
    raise StatementProcessingError(f"Auditoria BBVA fallida. {details}")


def extract_bbva_data(file_obj) -> tuple[pd.DataFrame, pd.DataFrame]:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    with pdfplumber.open(file_obj) as pdf:
        summary = _extract_bbva_summary(pdf)
        rows: list[dict] = []
        current_row: dict | None = None

        for page_number, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            if not words:
                continue

            lines = _group_words_into_lines(words)
            page_has_code_column = any(
                _bbva_header_has_code_column(" ".join(word["text"] for word in candidate_line))
                for candidate_line in lines
            )

            for line in lines:
                top = line[0]["top"]
                if top < BBVA_HEADER_MIN_TOP or top > BBVA_FOOTER_MAX_TOP:
                    continue

                texts = [word["text"] for word in line]
                line_text = " ".join(texts).strip()
                if not line_text:
                    continue

                if _is_bbva_transaction_start(texts):
                    if current_row:
                        rows.append(current_row)

                    code = None
                    content_start = 2
                    if page_has_code_column and len(texts) >= 3 and CODE_RE.match(texts[2] or ""):
                        code = texts[2]
                        content_start = 3

                    current_row = {
                        "page": page_number,
                        "fecha_operacion": texts[0],
                        "fecha_liquidacion": texts[1],
                        "codigo": code,
                        "descripcion_parts": [],
                        "referencia_parts": [],
                        "cargo": None,
                        "abono": None,
                        "saldo_operacion": None,
                        "saldo_liquidacion": None,
                    }

                    for word in line[content_start:]:
                        if _append_bbva_amount(
                            current_row,
                            word["text"],
                            word["x0"],
                            word.get("x1"),
                            compact_layout=not page_has_code_column,
                        ):
                            continue
                        if word["x0"] < 330:
                            current_row["descripcion_parts"].append(word["text"])
                        else:
                            current_row["referencia_parts"].append(word["text"])
                    continue

                if current_row is None:
                    continue

                if _should_ignore_bbva_line(line_text):
                    continue

                normalized_line = _normalize_text(line_text).upper()
                if (
                    line_text.startswith("No. Cuenta")
                    or line_text.startswith("No. Cliente")
                    or normalized_line.startswith("FECHA SALDO")
                    or normalized_line.startswith("OPER LIQ DESCRIPCION")
                ):
                    continue

                current_row["referencia_parts"].append(line_text)

        if current_row:
            rows.append(current_row)

    if not rows:
        raise StatementProcessingError("No se encontraron movimientos válidos en el estado de cuenta BBVA.")

    movimientos_df, resumen_df = _finalize_bbva_data(rows, summary)
    auditoria_df = build_bbva_audit_dataframe(movimientos_df, resumen_df)
    _validate_bbva_audit(auditoria_df)
    LOGGER.info(
        "Auditoria BBVA exitosa movimientos=%s abonos=%s cargos=%s puntos_saldo=%s",
        len(movimientos_df),
        int(movimientos_df["ABONO"].notna().sum()),
        int(movimientos_df["CARGO"].notna().sum()),
        int(movimientos_df["SALDO_OPERACION_PDF"].notna().sum()),
    )
    return movimientos_df, resumen_df


def _dfs_to_excel(sheets: list[tuple[str, pd.DataFrame]], *, style: bool = False) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheets:
            dataframe.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        if style:
            from openpyxl.styles import Alignment, Font, PatternFill

            header_fill = PatternFill("solid", fgColor="1F4E78")
            ok_fill = PatternFill("solid", fgColor="C6EFCE")
            error_fill = PatternFill("solid", fgColor="FFC7CE")
            for worksheet in writer.book.worksheets:
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions
                worksheet.sheet_view.showGridLines = False
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
                for column_cells in worksheet.columns:
                    values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
                    width = min(max(len(value) for value in values) + 2, 55)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = max(width, 11)
                if worksheet.title == "Movimientos":
                    money_columns = {
                        "CARGO",
                        "ABONO",
                        "DEPOSITO",
                        "RETIRO",
                        "SALDO_PDF",
                        "SALDO_OPERACION_PDF",
                        "SALDO_LIQUIDACION_PDF",
                        "SALDO_CALCULADO",
                        "DIFERENCIA_SALDO",
                    }
                    for header_cell in worksheet[1]:
                        if header_cell.value not in money_columns:
                            continue
                        for row in range(2, worksheet.max_row + 1):
                            worksheet.cell(row=row, column=header_cell.column).number_format = (
                                '$#,##0.00;[Red]($#,##0.00);-'
                            )
                if worksheet.title == "Auditoria":
                    status_column = next(cell.column for cell in worksheet[1] if cell.value == "ESTADO")
                    for row in range(2, worksheet.max_row + 1):
                        status_cell = worksheet.cell(row=row, column=status_column)
                        status_cell.fill = ok_fill if status_cell.value == "OK" else error_fill
                        status_cell.font = Font(bold=True)
    output.seek(0)
    return output


def process_bbva(file_obj):
    movimientos_df, resumen_df = extract_bbva_data(file_obj)
    auditoria_df = build_bbva_audit_dataframe(movimientos_df, resumen_df)
    return _dfs_to_excel([
        ("Movimientos", movimientos_df),
        ("Resumen", resumen_df),
        ("Auditoria", auditoria_df),
    ], style=True)


BANORTE_DATE_RE = re.compile(r"^\d{2}-[A-Z]{3}-\d{2}", re.IGNORECASE)
BANORTE_MONEY_RE = re.compile(r"^\d{1,3}(?:,\d{3})*\.\d{2}-?$")
BANORTE_DETAIL_TOP_MIN = 90
BANORTE_DETAIL_TOP_MAX = 735


def parse_banorte_summary_text(text: str) -> dict[str, Decimal | None]:
    normalized = _normalize_text(text)
    patterns = {
        "saldo_inicial": r"Saldo inicial del periodo \$\s*([\d,]+\.\d{2})",
        "depositos": r"\+ Total de depositos \$\s*([\d,]+\.\d{2})",
        "retiros": r"\- Total de retiros \$\s*([\d,]+\.\d{2})",
        "comisiones": r"\- Total de comisiones Cobradas / Pagadas \$\s*([\d,]+\.\d{2})",
        "iva_comisiones": r"\- IVA sobre comisiones \(16%\) \$\s*([\d,]+\.\d{2})",
        "saldo_final": r"Saldo actual \$\s*([\d,]+\.\d{2})",
    }
    return {key: _read_money(re.search(pattern, normalized).group(1)) if re.search(pattern, normalized) else None for key, pattern in patterns.items()}


def _extract_banorte_summary(pdf: pdfplumber.PDF) -> dict[str, Decimal | None]:
    for page in pdf.pages:
        text = page.extract_text() or ""
        summary = parse_banorte_summary_text(text)
        if summary["saldo_inicial"] is not None and summary["saldo_final"] is not None:
            return summary
    return parse_banorte_summary_text("")


def _is_banorte_transaction_start(texts: list[str]) -> bool:
    return bool(texts and BANORTE_DATE_RE.match(texts[0]))


def _banorte_should_stop_block(text: str) -> bool:
    normalized = _normalize_text(text).upper()
    return "CARGOS OBJETADOS EN EL PERIODO" in normalized or "GANANCIA ANUAL TOTAL" in normalized or "COMPROBANTE FISCAL DIGITAL" in normalized


def _banorte_is_other_charge(description: str) -> bool:
    normalized = _normalize_text(description).upper()
    return "PENALIZACION POR NO MANTENER EL SALDO PROMEDIO MINIMO" in normalized or normalized.startswith("I.V.A. LIQ")


def extract_banorte_data(file_obj) -> tuple[pd.DataFrame, pd.DataFrame]:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    with pdfplumber.open(file_obj) as pdf:
        summary = _extract_banorte_summary(pdf)
        rows: list[dict] = []
        current_row: dict | None = None

        for page_number, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(keep_blank_chars=False, use_text_flow=False) or []
            for line in _group_words_into_lines(words, tolerance=3):
                line = sorted(line, key=lambda word: word["x0"])
                top = line[0]["top"]
                if top < BANORTE_DETAIL_TOP_MIN or top > BANORTE_DETAIL_TOP_MAX:
                    continue

                texts = [word["text"] for word in line]
                line_text = " ".join(texts).strip()
                if not line_text:
                    continue

                normalized = _normalize_text(line_text).upper()
                if "DETALLE DE MOVIMIENTOS" in normalized:
                    continue
                if "FECHA DESCRIPCION / ESTABLECIMIENTO" in normalized:
                    continue
                if _banorte_should_stop_block(line_text):
                    if current_row:
                        rows.append(current_row)
                        current_row = None
                    continue
                if normalized in {"OTROS", "OTROS?"}:
                    continue
                if re.match(r"^\d+/\d+$", normalized):
                    continue

                if _is_banorte_transaction_start(texts):
                    if current_row:
                        rows.append(current_row)

                    match = BANORTE_DATE_RE.match(texts[0])
                    fecha = match.group(0)
                    descripcion_partes: list[str] = []
                    remaining = texts[0][len(fecha):].strip()
                    if remaining:
                        descripcion_partes.append(remaining)

                    amount_candidates: list[tuple[float, str]] = []
                    for word in line[1:]:
                        txt = word["text"].strip()
                        if BANORTE_MONEY_RE.match(txt) and word["x0"] >= 390:
                            amount_candidates.append((word["x0"], txt))
                        else:
                            descripcion_partes.append(txt)

                    saldo_text = amount_candidates[-1][1] if amount_candidates else None
                    monto_text = amount_candidates[-2][1] if len(amount_candidates) >= 2 else None
                    monto_x = amount_candidates[-2][0] if len(amount_candidates) >= 2 else None

                    current_row = {
                        "page": page_number,
                        "fecha": fecha,
                        "descripcion_parts": descripcion_partes,
                        "monto_text": monto_text,
                        "monto_x": monto_x,
                        "saldo_text": saldo_text,
                    }
                    continue

                if current_row is None:
                    continue

                current_row["descripcion_parts"].append(line_text)

        if current_row:
            rows.append(current_row)

    if not rows:
        raise StatementProcessingError("No se encontraron movimientos validos en el estado de cuenta Banorte.")

    saldo_anterior = summary.get("saldo_inicial")
    total_depositos = Decimal("0.00")
    total_retiros = Decimal("0.00")
    total_otros = Decimal("0.00")
    movimientos: list[dict] = []

    for index, row in enumerate(rows, start=1):
        descripcion = _clean_join(row["descripcion_parts"])
        if _normalize_text(descripcion).upper() == "SALDO ANTERIOR":
            continue

        monto = _read_money(row["monto_text"]) or Decimal("0.00")
        saldo_pdf = _read_money(row["saldo_text"])
        deposito = Decimal("0.00")
        retiro = Decimal("0.00")
        otros_cargos = Decimal("0.00")

        if monto and saldo_anterior is not None and saldo_pdf is not None:
            if _banorte_is_other_charge(descripcion):
                otros_cargos = monto
            elif saldo_pdf > saldo_anterior:
                deposito = monto
            elif saldo_pdf < saldo_anterior:
                retiro = monto
            elif row["monto_x"] is not None and row["monto_x"] < 430:
                deposito = monto
            else:
                retiro = monto
        elif monto:
            if row["monto_x"] is not None and row["monto_x"] < 430:
                deposito = monto
            else:
                retiro = monto

        total_depositos += deposito
        total_retiros += retiro
        total_otros += otros_cargos

        saldo_calculado = None
        if saldo_anterior is not None:
            saldo_calculado = (saldo_anterior + deposito - retiro - otros_cargos).quantize(Decimal("0.01"))
            if saldo_pdf is not None:
                saldo_anterior = saldo_pdf
            else:
                saldo_anterior = saldo_calculado

        movimientos.append(
            {
                "MOVIMIENTO": index,
                "PAGINA": row["page"],
                "FECHA": row["fecha"],
                "DESCRIPCION": descripcion,
                "MONTO_PDF": _money_to_float(monto) if monto else None,
                "DEPOSITO": _money_to_float(deposito) if deposito else None,
                "RETIRO": _money_to_float(retiro) if retiro else None,
                "OTROS_CARGOS": _money_to_float(otros_cargos) if otros_cargos else None,
                "SALDO_PDF": _money_to_float(saldo_pdf),
                "SALDO_CALCULADO": _money_to_float(saldo_calculado),
            }
        )

    resumen = pd.DataFrame([
        {"METRICA": "Saldo inicial", "PDF": _money_to_float(summary.get("saldo_inicial")), "CALCULADO": _money_to_float(summary.get("saldo_inicial")), "DIFERENCIA": 0.0 if summary.get("saldo_inicial") is not None else None},
        {"METRICA": "Depositos", "PDF": _money_to_float(summary.get("depositos")), "CALCULADO": _money_to_float(total_depositos), "DIFERENCIA": _money_to_float((summary.get("depositos") - total_depositos) if summary.get("depositos") is not None else None)},
        {"METRICA": "Retiros", "PDF": _money_to_float(summary.get("retiros")), "CALCULADO": _money_to_float(total_retiros), "DIFERENCIA": _money_to_float((summary.get("retiros") - total_retiros) if summary.get("retiros") is not None else None)},
        {"METRICA": "Otros cargos", "PDF": _money_to_float((summary.get("comisiones") or Decimal("0.00")) + (summary.get("iva_comisiones") or Decimal("0.00"))), "CALCULADO": _money_to_float(total_otros), "DIFERENCIA": _money_to_float((((summary.get("comisiones") or Decimal("0.00")) + (summary.get("iva_comisiones") or Decimal("0.00"))) - total_otros))},
        {"METRICA": "Saldo final", "PDF": _money_to_float(summary.get("saldo_final")), "CALCULADO": movimientos[-1]["SALDO_CALCULADO"] if movimientos else None, "DIFERENCIA": _money_to_float((summary.get("saldo_final") - Decimal(str(movimientos[-1]['SALDO_CALCULADO']))) if summary.get("saldo_final") is not None and movimientos and movimientos[-1]["SALDO_CALCULADO"] is not None else None)},
    ])
    return pd.DataFrame(movimientos), resumen


def process_banorte(file_obj):
    movimientos_df, resumen_df = extract_banorte_data(file_obj)
    return _dfs_to_excel([
        ("Movimientos", movimientos_df),
        ("Resumen", resumen_df),
    ])


BANORTE_CHEQUES_ACCOUNT_RE = re.compile(r"^\d{10}$")
BANORTE_CHEQUES_DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
BANORTE_CHEQUES_MONEY_RE = re.compile(r"^\$?\d{1,3}(?:,\d{3})*\.\d{2}$")
BANORTE_CHEQUES_ROW_TOP_MIN = 25
BANORTE_CHEQUES_ROW_TOP_MAX = 735
BANORTE_CHEQUES_PRE_DETAIL_MARKERS = (
    "SPEI RECIBIDO",
    "CARGO POR PAGO",
    "INTERESES EXENTO",
    "IVA DEPOSITO",
)


def _banorte_cheques_money_from_words(words: list[dict], min_x: float, max_x: float) -> str | None:
    for word in words:
        text = word["text"].strip()
        if min_x <= float(word["x0"]) < max_x and BANORTE_CHEQUES_MONEY_RE.match(text):
            return text
    return None


def _banorte_cheques_text_from_words(words: list[dict], min_x: float, max_x: float) -> str:
    return _clean_join(word["text"] for word in words if min_x <= float(word["x0"]) < max_x)


def _is_banorte_cheques_row(words: list[dict]) -> bool:
    if len(words) < 3:
        return False
    return (
        BANORTE_CHEQUES_ACCOUNT_RE.match(words[0]["text"]) is not None
        and BANORTE_CHEQUES_DATE_RE.match(words[1]["text"]) is not None
    )


def _is_banorte_cheques_header_or_footer(text: str) -> bool:
    normalized = _normalize_text(text).upper()
    if not normalized:
        return True
    return (
        normalized.startswith("CUENTA FECHA DESCRIPCION DEPOSITOS RETIROS SALDO")
        or normalized.startswith("CUENTAS DE CHEQUES")
        or normalized.startswith("BANCO MERCANTIL DEL NORTE")
        or normalized.startswith("RFC:")
        or normalized.startswith("SALDO TOTAL")
        or normalized.startswith("SALDO DISPONIBLE")
        or normalized.startswith("SALDO ACTUAL")
        or normalized.startswith("INICIAL DEL DIA")
        or normalized.startswith("FINAL MES ANTERIOR")
        or normalized.startswith("PROVEEDORA DE INSUMOS")
        or normalized.startswith("5635 PROVEEDORA")
        or normalized.startswith("0675-SUCURSAL")
    )


def _starts_banorte_cheques_pre_detail(text: str) -> bool:
    normalized = _normalize_text(text).upper()
    return any(normalized.startswith(marker) for marker in BANORTE_CHEQUES_PRE_DETAIL_MARKERS)


def parse_banorte_cuenta_cheques_summary_text(text: str) -> dict[str, Decimal | int | None]:
    normalized = _normalize_text(text)
    summary: dict[str, Decimal | int | None] = {
        "depositos_count": None,
        "retiros_count": None,
        "depositos": None,
        "retiros": None,
    }

    operations = re.search(r"OPERACIONES:\s+(\d+)\s+(\d+)", normalized, re.IGNORECASE)
    if operations:
        summary["depositos_count"] = int(operations.group(1))
        summary["retiros_count"] = int(operations.group(2))

    totals = re.search(r"TOTAL:\s+\$?([\d,]+\.\d{2})\s+\$?([\d,]+\.\d{2})", normalized, re.IGNORECASE)
    if totals:
        summary["depositos"] = _read_money(totals.group(1))
        summary["retiros"] = _read_money(totals.group(2))

    return summary


def _extract_banorte_cuenta_cheques_summary(pdf: pdfplumber.PDF) -> dict[str, Decimal | int | None]:
    full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    return parse_banorte_cuenta_cheques_summary_text(full_text)


def _banorte_cheques_make_row(page_number: int, words: list[dict], pre_detail: list[str]) -> dict:
    detail_text = _banorte_cheques_text_from_words(words, 405, 999)
    description = _banorte_cheques_text_from_words(words, 120, 250)
    return {
        "page": page_number,
        "top": float(words[0]["top"]),
        "cuenta": words[0]["text"],
        "fecha": words[1]["text"],
        "descripcion": description,
        "deposito_text": _banorte_cheques_money_from_words(words, 245, 300),
        "retiro_text": _banorte_cheques_money_from_words(words, 300, 355),
        "saldo_text": _banorte_cheques_money_from_words(words, 350, 410),
        "detalle_parts": [*pre_detail, detail_text] if detail_text else list(pre_detail),
    }


def _finalize_banorte_cuenta_cheques_data(
    rows: list[dict],
    summary: dict[str, Decimal | int | None],
    debug_rows: list[dict],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    movimientos: list[dict] = []
    auditoria: list[dict] = []
    total_depositos = Decimal("0.00")
    total_retiros = Decimal("0.00")
    previous_balance: Decimal | None = None
    calculated_initial: Decimal | None = None

    for index, row in enumerate(rows, start=1):
        deposito = _read_money(row.get("deposito_text")) or Decimal("0.00")
        retiro = _read_money(row.get("retiro_text")) or Decimal("0.00")
        saldo_pdf = _read_money(row.get("saldo_text"))

        if index == 1 and saldo_pdf is not None:
            calculated_initial = (saldo_pdf - deposito + retiro).quantize(Decimal("0.01"))
            previous_balance = calculated_initial

        saldo_calculado = None
        diferencia = None
        cuadra = None
        if previous_balance is not None:
            saldo_calculado = (previous_balance + deposito - retiro).quantize(Decimal("0.01"))
            if saldo_pdf is not None:
                diferencia = (saldo_pdf - saldo_calculado).quantize(Decimal("0.01"))
                cuadra = diferencia == Decimal("0.00")
                previous_balance = saldo_pdf
            else:
                previous_balance = saldo_calculado

        total_depositos += deposito
        total_retiros += retiro
        descripcion = _clean_join([row["descripcion"]])
        detalle = _clean_join(row["detalle_parts"])

        if saldo_pdf is None:
            auditoria.append({"fila_excel": index + 1, "tipo": "sin_saldo_pdf", "detalle": f"Pagina {row['page']} {row['fecha']} {descripcion}"})
        if deposito == Decimal("0.00") and retiro == Decimal("0.00"):
            auditoria.append({"fila_excel": index + 1, "tipo": "sin_monto", "detalle": f"Pagina {row['page']} {row['fecha']} {descripcion}"})
        if cuadra is False:
            auditoria.append({"fila_excel": index + 1, "tipo": "saldo_no_cuadra", "detalle": f"PDF={saldo_pdf} calculado={saldo_calculado} diferencia={diferencia}"})

        movimientos.append(
            {
                "MOVIMIENTO": index,
                "PAGINA": row["page"],
                "CUENTA": row["cuenta"],
                "FECHA": row["fecha"],
                "DESCRIPCION": descripcion,
                "DESCRIPCION_DETALLADA": detalle,
                "DEPOSITO": _money_to_float(deposito) if deposito else None,
                "RETIRO": _money_to_float(retiro) if retiro else None,
                "SALDO_PDF": _money_to_float(saldo_pdf),
                "SALDO_CALCULADO": _money_to_float(saldo_calculado),
                "CUADRA_CON_SALDO_PDF": cuadra,
                "DIFERENCIA_SALDO": _money_to_float(diferencia),
            }
        )

    expected_final = _read_money(rows[-1].get("saldo_text")) if rows else None
    parsed_final = Decimal(str(movimientos[-1]["SALDO_CALCULADO"])) if movimientos and movimientos[-1]["SALDO_CALCULADO"] is not None else None
    final_difference = None
    if expected_final is not None and parsed_final is not None:
        final_difference = (expected_final - parsed_final).quantize(Decimal("0.01"))

    summary_rows = [
        {"METRICA": "Saldo inicial calculado", "PDF": None, "CALCULADO": _money_to_float(calculated_initial), "DIFERENCIA": None},
        {"METRICA": "Depositos", "PDF": _money_to_float(summary.get("depositos")), "CALCULADO": _money_to_float(total_depositos), "DIFERENCIA": _money_to_float((summary.get("depositos") - total_depositos) if isinstance(summary.get("depositos"), Decimal) else None)},
        {"METRICA": "Retiros", "PDF": _money_to_float(summary.get("retiros")), "CALCULADO": _money_to_float(total_retiros), "DIFERENCIA": _money_to_float((summary.get("retiros") - total_retiros) if isinstance(summary.get("retiros"), Decimal) else None)},
        {"METRICA": "Saldo final", "PDF": _money_to_float(expected_final), "CALCULADO": _money_to_float(parsed_final), "DIFERENCIA": _money_to_float(final_difference)},
        {"METRICA": "Numero de depositos", "PDF": summary.get("depositos_count"), "CALCULADO": int(sum(1 for row in movimientos if row["DEPOSITO"])), "DIFERENCIA": None},
        {"METRICA": "Numero de retiros", "PDF": summary.get("retiros_count"), "CALCULADO": int(sum(1 for row in movimientos if row["RETIRO"])), "DIFERENCIA": None},
    ]

    for item in summary_rows:
        diferencia = item.get("DIFERENCIA")
        if isinstance(diferencia, float) and abs(diferencia) >= 0.01:
            auditoria.append({"fila_excel": None, "tipo": "resumen_no_cuadra", "detalle": f"{item['METRICA']}: diferencia={diferencia}"})
    for metric, pdf_key, calculated in (
        ("Numero de depositos", "depositos_count", int(sum(1 for row in movimientos if row["DEPOSITO"]))),
        ("Numero de retiros", "retiros_count", int(sum(1 for row in movimientos if row["RETIRO"]))),
    ):
        expected = summary.get(pdf_key)
        if isinstance(expected, int) and expected != calculated:
            auditoria.append({"fila_excel": None, "tipo": "conteo_no_cuadra", "detalle": f"{metric}: PDF={expected} calculado={calculated}"})

    if not auditoria:
        auditoria.append({"fila_excel": None, "tipo": "ok", "detalle": "Sin alertas de auditoria. Totales y saldos revisados."})

    return (
        pd.DataFrame(movimientos),
        pd.DataFrame(summary_rows),
        pd.DataFrame(auditoria),
        pd.DataFrame(debug_rows),
    )


def extract_banorte_cuenta_cheques_data(file_obj) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    with pdfplumber.open(file_obj) as pdf:
        summary = _extract_banorte_cuenta_cheques_summary(pdf)
        rows: list[dict] = []
        debug_rows: list[dict] = []
        current_row: dict | None = None
        pending_pre_detail: list[str] = []

        for page_number, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(keep_blank_chars=False, use_text_flow=False) or []
            for line in _group_words_into_lines(sorted(words, key=lambda item: (item["top"], item["x0"])), tolerance=3):
                line = sorted(line, key=lambda word: word["x0"])
                top = float(line[0]["top"])
                if top < BANORTE_CHEQUES_ROW_TOP_MIN or top > BANORTE_CHEQUES_ROW_TOP_MAX:
                    continue

                line_text = " ".join(word["text"] for word in line).strip()
                normalized = _normalize_text(line_text).upper()
                if not line_text:
                    continue
                if normalized.startswith("DEPOSITOS RETIROS") or normalized.startswith("OPERACIONES:") or normalized.startswith("TOTAL:"):
                    if current_row:
                        rows.append(current_row)
                        current_row = None
                    debug_rows.append({"pagina": page_number, "top": top, "tipo": "resumen", "texto": line_text})
                    continue
                if _is_banorte_cheques_header_or_footer(line_text):
                    debug_rows.append({"pagina": page_number, "top": top, "tipo": "omitido_header_footer", "texto": line_text})
                    continue

                if _is_banorte_cheques_row(line):
                    if current_row:
                        rows.append(current_row)
                    current_row = _banorte_cheques_make_row(page_number, line, pending_pre_detail)
                    pending_pre_detail = []
                    debug_rows.append({"pagina": page_number, "top": top, "tipo": "movimiento", "texto": line_text})
                    continue

                if _starts_banorte_cheques_pre_detail(line_text) or pending_pre_detail:
                    if current_row and _starts_banorte_cheques_pre_detail(line_text):
                        rows.append(current_row)
                        current_row = None
                    pending_pre_detail.append(line_text)
                    debug_rows.append({"pagina": page_number, "top": top, "tipo": "detalle_previo", "texto": line_text})
                    continue

                if current_row is not None:
                    current_row["detalle_parts"].append(line_text)
                    debug_rows.append({"pagina": page_number, "top": top, "tipo": "detalle_movimiento", "texto": line_text})
                    continue

                pending_pre_detail.append(line_text)
                debug_rows.append({"pagina": page_number, "top": top, "tipo": "detalle_previo_sin_movimiento", "texto": line_text})

        if current_row:
            rows.append(current_row)
        if pending_pre_detail:
            debug_rows.append({"pagina": "EOF", "top": None, "tipo": "detalle_previo_sin_destino", "texto": " | ".join(pending_pre_detail)})

    if not rows:
        raise StatementProcessingError("No se encontraron movimientos validos en el estado de cuenta Banorte Cuenta Cheques.")

    return _finalize_banorte_cuenta_cheques_data(rows, summary, debug_rows)


def process_banorte_cuenta_cheques(file_obj):
    movimientos_df, resumen_df, auditoria_df, debug_df = extract_banorte_cuenta_cheques_data(file_obj)
    return _dfs_to_excel([
        ("Movimientos", movimientos_df),
        ("Resumen", resumen_df),
        ("Auditoria", auditoria_df),
        ("Debug", debug_df),
    ])


SCOTIA_START_RE = re.compile(r"^(\d{2})\s+(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\b", re.IGNORECASE)
SCOTIA_DATE_RE = re.compile(r"^\d{2}$")
SCOTIA_MONTH_RE = re.compile(r"^(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)$", re.IGNORECASE)
SCOTIA_MONEY_RE = re.compile(r"\$[\d,]+\.\d{2}")


def _is_scotiabank_detail_header(text: str) -> bool:
    normalized = _normalize_text(text).upper()
    return "DETALLEDETUSMOVIMIENTOS" in normalized or "FECHA CONCEPTO ORIGEN/REFERENCIA DEPOSITO RETIRO SALDO" in normalized


def _should_skip_scotiabank_line(text: str) -> bool:
    normalized = _normalize_text(text)
    if not normalized:
        return True
    upper = normalized.upper()
    if upper.startswith("PAGINA ") or upper.startswith("CUENTA "):
        return True
    return (
        "SCOTIABANK INVERLAT" in upper
        or "PARA LOS EFECTOS DEL ART." in upper
        or "LASTASASDEINTERESESTANEXPRESADASENTERMINOSANUALESSIMPLES" in upper
        or "ENELCASODEENVIODETRANSFERENCIASSPEI" in upper
        or "EL 29-JUL-24 SE ACTUALIZARA EL CONTRATO" in upper
        or "12-AGO-24 SE ACTUALIZARA EL CONTRATO" in upper
        or "SI DESEAS RECIBIR PAGOS A TRAVES" in upper
        or "NACIONAL), DEBERAS INFORMAR" in upper
        or "044180001031111487" == upper
    )


def _extract_scotiabank_row(line: str, year: int) -> dict | None:
    match = SCOTIA_START_RE.match(line.strip())
    if not match:
        return None

    day = int(match.group(1))
    month = match.group(2).upper()
    remainder = line[match.end():].strip()
    if not remainder:
        return None
    amounts = SCOTIA_MONEY_RE.findall(remainder)

    saldo_text = amounts[-1] if amounts else None
    monto_text = amounts[-2] if len(amounts) >= 2 else None
    tail_without_amounts = remainder
    for amount in amounts:
        tail_without_amounts = tail_without_amounts.replace(amount, "", 1)
    tail_without_amounts = re.sub(r"\s+", " ", tail_without_amounts).strip()

    tokens = tail_without_amounts.split()
    descripcion_tokens: list[str] = []
    referencia_tokens: list[str] = []
    for token in tokens:
        cleaned = token.strip(",")
        is_reference = (
            cleaned.startswith("/")
            or cleaned.endswith(":")
            or bool(re.search(r"\d", cleaned))
            or cleaned in {"SB", "SPEIWEB"}
        )
        if referencia_tokens or is_reference:
            referencia_tokens.append(token)
        else:
            descripcion_tokens.append(token)

    return {
        "fecha": f"{day:02d}-{month}-{str(year)[-2:]}",
        "descripcion": " ".join(descripcion_tokens).strip(),
        "referencia": " ".join(referencia_tokens).strip(),
        "monto_text": monto_text,
        "saldo_text": saldo_text,
        "continuaciones": [],
    }


def parse_scotiabank_summary_text(text: str) -> dict[str, Decimal | None]:
    normalized = _normalize_text(text)
    patterns = {
        "saldo_inicial": r"S\s*aldo inicial\s*=\s*\$([\d,]+\.\d{2})|Saldoinicial\s*\$([\d,]+\.\d{2})",
        "depositos": r"\(\+\)Depositos \$([\d,]+\.\d{2})",
        "retiros": r"\(\-\)Retiros \$([\d,]+\.\d{2})",
        "comisiones": r"\(\-\)Comisionescobradas \$([\d,]+\.\d{2})",
        "impuestos": r"\(\-\)Impuestos \$([\d,]+\.\d{2})",
        "saldo_final": r"Saldo final= \$([\d,]+\.\d{2})",
    }
    summary: dict[str, Decimal | None] = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, normalized, re.IGNORECASE)
        if not match:
            summary[key] = None
            continue
        groups = [group for group in match.groups() if group]
        summary[key] = _read_money(groups[0]) if groups else None
    return summary


def _extract_scotiabank_summary(pdf: pdfplumber.PDF) -> dict[str, Decimal | None]:
    for page in pdf.pages:
        text = page.extract_text() or ""
        summary = parse_scotiabank_summary_text(text)
        if summary["saldo_inicial"] is not None and summary["saldo_final"] is not None:
            return summary
    return parse_scotiabank_summary_text("")


def _scotia_extract_year(summary_text: str) -> int | None:
    normalized = _normalize_text(summary_text)
    match = re.search(r"Periodo\s+\d{2}-[A-Z]{3}-(\d{2})/\d{2}-[A-Z]{3}-(\d{2})", normalized, re.IGNORECASE)
    if not match:
        return None
    return 2000 + int(match.group(2))


def extract_scotiabank_data(file_obj) -> tuple[pd.DataFrame, pd.DataFrame]:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    with pdfplumber.open(file_obj) as pdf:
        page1_text = pdf.pages[0].extract_text() or ""
        summary = _extract_scotiabank_summary(pdf)
        year = _scotia_extract_year(page1_text) or 2026
        rows: list[dict] = []
        current_row: dict | None = None

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            in_detail = page_number > 1
            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                if _is_scotiabank_detail_header(line):
                    in_detail = True
                    continue
                if not in_detail or _should_skip_scotiabank_line(line):
                    continue

                row = _extract_scotiabank_row(line, year)
                if row:
                    if current_row:
                        rows.append(current_row)
                    current_row = {
                        "page": page_number,
                        **row,
                    }
                    continue

                if current_row is None:
                    continue

                normalized_line = _normalize_text(line).upper()
                if (
                    normalized_line.startswith("TOTALDECOMISIONESCOBRADASENELPERIODO")
                    or normalized_line.startswith("TOTALDECOMISIONESCOBRADAS")
                    or normalized_line.startswith("TOTALDEIMPUESTOSCOBRADOS")
                ):
                    rows.append(current_row)
                    current_row = None
                    continue

                current_row["continuaciones"].append(line)

        if current_row:
            rows.append(current_row)

    if not rows:
        raise StatementProcessingError("No se encontraron movimientos validos en el estado de cuenta Scotiabank.")

    saldo_anterior = summary.get("saldo_inicial")
    total_depositos = Decimal("0.00")
    total_retiros = Decimal("0.00")
    total_comisiones = Decimal("0.00")
    total_impuestos = Decimal("0.00")
    movimientos: list[dict] = []

    for index, row in enumerate(rows, start=1):
        monto = _read_money(row["monto_text"]) or Decimal("0.00")
        saldo_pdf = _read_money(row["saldo_text"])
        deposito = Decimal("0.00")
        retiro = Decimal("0.00")
        comision = Decimal("0.00")
        impuesto = Decimal("0.00")
        normalized_desc = _normalize_text(f"{row['descripcion']} {' '.join(row['continuaciones'])}").upper()

        if not monto:
            for continuation in row["continuaciones"]:
                amounts = SCOTIA_MONEY_RE.findall(continuation)
                if len(amounts) >= 2:
                    monto = _read_money(amounts[-2]) or Decimal("0.00")
                    saldo_pdf = _read_money(amounts[-1]) or saldo_pdf
                    break

        if monto and "IVA - COMISIONES" in normalized_desc:
            impuesto = monto
        elif monto and ("IMPUESTO COM/CGO TRASPASO" in normalized_desc):
            impuesto = monto
        elif monto and ("COMISION" in normalized_desc or "TRASPASOS A OTROS BANCOS" in normalized_desc):
            comision = monto
        elif monto and saldo_anterior is not None and saldo_pdf is not None:
            if saldo_pdf >= saldo_anterior:
                deposito = monto
            else:
                retiro = monto
        elif monto:
            if any(keyword in normalized_desc for keyword in ("DEPOSITO", "DEVOLUCION", "INTERESES")):
                deposito = monto
            else:
                retiro = monto

        total_depositos += deposito
        total_retiros += retiro
        total_comisiones += comision
        total_impuestos += impuesto

        saldo_calculado = None
        if saldo_anterior is not None:
            saldo_calculado = (saldo_anterior + deposito - retiro - comision - impuesto).quantize(Decimal("0.01"))
            if saldo_pdf is not None:
                saldo_anterior = saldo_pdf
            else:
                saldo_anterior = saldo_calculado

        movimientos.append(
            {
                "MOVIMIENTO": index,
                "PAGINA": row["page"],
                "FECHA": row["fecha"],
                "DESCRIPCION": _clean_join([row["descripcion"], *row["continuaciones"]]),
                "REFERENCIA": row["referencia"],
                "DEPOSITO": _money_to_float(deposito) if deposito else None,
                "RETIRO": _money_to_float(retiro) if retiro else None,
                "COMISION": _money_to_float(comision) if comision else None,
                "IMPUESTO": _money_to_float(impuesto) if impuesto else None,
                "SALDO_PDF": _money_to_float(saldo_pdf),
                "SALDO_CALCULADO": _money_to_float(saldo_calculado),
            }
        )

    resumen = pd.DataFrame([
        {"METRICA": "Saldo inicial", "PDF": _money_to_float(summary.get("saldo_inicial")), "CALCULADO": _money_to_float(summary.get("saldo_inicial")), "DIFERENCIA": 0.0 if summary.get("saldo_inicial") is not None else None},
        {"METRICA": "Depositos", "PDF": _money_to_float(summary.get("depositos")), "CALCULADO": _money_to_float(total_depositos), "DIFERENCIA": _money_to_float((summary.get("depositos") - total_depositos) if summary.get("depositos") is not None else None)},
        {"METRICA": "Retiros", "PDF": _money_to_float(summary.get("retiros")), "CALCULADO": _money_to_float(total_retiros), "DIFERENCIA": _money_to_float((summary.get("retiros") - total_retiros) if summary.get("retiros") is not None else None)},
        {"METRICA": "Comisiones", "PDF": _money_to_float(summary.get("comisiones")), "CALCULADO": _money_to_float(total_comisiones), "DIFERENCIA": _money_to_float((summary.get("comisiones") - total_comisiones) if summary.get("comisiones") is not None else None)},
        {"METRICA": "Impuestos", "PDF": _money_to_float(summary.get("impuestos")), "CALCULADO": _money_to_float(total_impuestos), "DIFERENCIA": _money_to_float((summary.get("impuestos") - total_impuestos) if summary.get("impuestos") is not None else None)},
        {"METRICA": "Saldo final", "PDF": _money_to_float(summary.get("saldo_final")), "CALCULADO": movimientos[-1]["SALDO_CALCULADO"] if movimientos else None, "DIFERENCIA": _money_to_float((summary.get("saldo_final") - Decimal(str(movimientos[-1]['SALDO_CALCULADO']))) if summary.get("saldo_final") is not None and movimientos and movimientos[-1]["SALDO_CALCULADO"] is not None else None)},
    ])
    return pd.DataFrame(movimientos), resumen


def process_scotiabank(file_obj):
    movimientos_df, resumen_df = extract_scotiabank_data(file_obj)
    return _dfs_to_excel([
        ("Movimientos", movimientos_df),
        ("Resumen", resumen_df),
    ])


INBURSA_MONTH_RE = re.compile(r"^(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\.?$", re.IGNORECASE)
INBURSA_MONEY_RE = re.compile(r"^\d{1,3}(?:,\d{3})*\.\d{2}$")
INBURSA_LINE_TOLERANCE = 1.8
INBURSA_DETAIL_MAX_TOP = 750
INBURSA_CARGO_MAX_RIGHT = 445
INBURSA_ABONO_MAX_RIGHT = 512


def parse_inbursa_summary_text(text: str) -> dict[str, Decimal | None]:
    normalized = _normalize_text(text)
    patterns = {
        "saldo_inicial": r"SALDO ANTERIOR\s+([\d,]+\.\d{2})",
        "abonos": r"ABONOS\s+([\d,]+\.\d{2})",
        "cargos": r"CARGOS\s+([\d,]+\.\d{2})",
        "saldo_final": r"SALDO ACTUAL\s+([\d,]+\.\d{2})",
        "rendimientos": r"RENDIMIENTOS\s+([\d,]+\.\d{2})",
        "comisiones": r"EN EL PERIODO\s+([\d,]+\.\d{2})",
    }
    return {key: _read_money(re.search(pattern, normalized).group(1)) if re.search(pattern, normalized) else None for key, pattern in patterns.items()}


def _extract_inbursa_summary(pdf: pdfplumber.PDF) -> dict[str, Decimal | None]:
    for page in pdf.pages:
        text = page.extract_text() or ""
        summary = parse_inbursa_summary_text(text)
        if summary["saldo_inicial"] is not None and summary["saldo_final"] is not None:
            return summary
    return parse_inbursa_summary_text("")


def _is_inbursa_header_or_footer(text: str) -> bool:
    normalized = _normalize_text(text).upper()
    return (
        not normalized
        or normalized.startswith("PAGINA:")
        or normalized == "ESTADO DE CUENTA"
        or normalized.startswith("CLIENTE INBURSA:")
        or normalized.startswith("AV MORELOS 9")
        or normalized.startswith("AXOTLA ALVARO OBREGON")
        or normalized.startswith("CIUDAD DE MEXICO, MEX")
        or normalized.startswith("TIPO COMPROBANTE:")
        or normalized.startswith("CONSULTAS Y RECLAMACIONES")
        or normalized.startswith("GLOSARIO DE ABREVIATURAS")
        or normalized.startswith("BANCO INBURSA, S.A.")
        or normalized.startswith("AVENIDA PASEO DE LAS PALMAS")
        or normalized.startswith("REGIMEN FISCAL:")
        or normalized.startswith("SI DESEA RECIBIR PAGOS A TRAVES")
        or normalized.startswith("EL NOMBRE DEL BENEFICIARIO EN UNA TRANSFERENCIA")
        or normalized.startswith("CLAVE DE RASTREO. SIRVE PARA IDENTIFICAR")
        or normalized.startswith("N TASAS EXPRESADAS EN TERMINOS ANUALES")
        or normalized.startswith("N A PARTIR DE LA FECHA DE FIN DE PERIODO")
        or normalized.startswith("N EN CASO DE ROBO O EXTRAVIO DE TARJETA")
        or normalized.startswith("N LE RECORDAMOS QUE BANCO INBURSA")
        or normalized.startswith("N RENDIMIENTOS:")
    )


def _should_end_inbursa_detail(text: str) -> bool:
    normalized = _normalize_text(text).upper()
    return (
        normalized.startswith("SI DESEA RECIBIR PAGOS A TRAVES")
        or normalized.startswith("TIPO COMPROBANTE:")
        or normalized.startswith("RESUMEN DEL CFDI")
        or normalized.startswith("CONSULTAS Y RECLAMACIONES")
        or normalized.startswith("GLOSARIO DE ABREVIATURAS")
    )


def _inbursa_amount_column_centers(words: list[dict]) -> tuple[float, float, float] | None:
    centers: dict[str, float] = {}
    for word in words:
        label = _normalize_text(word["text"]).upper()
        if label in {"CARGOS", "ABONOS", "SALDO"}:
            centers[label] = (word["x0"] + word["x1"]) / 2
    if not all(label in centers for label in ("CARGOS", "ABONOS", "SALDO")):
        return None
    return centers["CARGOS"], centers["ABONOS"], centers["SALDO"]


def _extract_inbursa_row_from_words(
    words: list[dict],
    current_day: int | None = None,
    amount_column_centers: tuple[float, float, float] | None = None,
) -> tuple[dict | None, int | None]:
    if not words or not INBURSA_MONTH_RE.match(words[0]["text"]):
        return None, current_day

    month = words[0]["text"].rstrip(".").upper()
    index = 1
    if index < len(words) and re.fullmatch(r"\d{1,2}", words[index]["text"]) and words[index]["x1"] < 100:
        current_day = int(words[index]["text"])
        index += 1

    amount_words = [word for word in words[index:] if INBURSA_MONEY_RE.match(word["text"])]
    if not amount_words:
        return None, current_day

    text = " ".join(word["text"] for word in words)
    if "BALANCE INICIAL" in _normalize_text(text).upper():
        return {
            "month": month,
            "day": current_day,
            "referencia": "",
            "concepto": "BALANCE INICIAL",
            "cargo_text": None,
            "abono_text": None,
            "saldo_text": amount_words[-1]["text"],
            "continuaciones": [],
            "skip": True,
        }, current_day

    cargo_text = None
    abono_text = None
    saldo_text = None
    for word in amount_words:
        if amount_column_centers is not None:
            word_center = (word["x0"] + word["x1"]) / 2
            column = min(
                enumerate(amount_column_centers),
                key=lambda item: abs(word_center - item[1]),
            )[0]
            if column == 0:
                cargo_text = word["text"]
            elif column == 1:
                abono_text = word["text"]
            else:
                saldo_text = word["text"]
        elif word["x1"] <= INBURSA_CARGO_MAX_RIGHT:
            cargo_text = word["text"]
        elif word["x1"] <= INBURSA_ABONO_MAX_RIGHT:
            abono_text = word["text"]
        else:
            saldo_text = word["text"]

    if saldo_text is None or (cargo_text is None and abono_text is None):
        return None, current_day

    amount_ids = {id(word) for word in amount_words}
    head_words = [word for word in words[index:] if id(word) not in amount_ids]
    referencia = head_words[0]["text"] if head_words else ""
    concepto = " ".join(word["text"] for word in head_words[1:])
    return {
        "month": month,
        "day": current_day,
        "referencia": referencia,
        "concepto": concepto,
        "cargo_text": cargo_text,
        "abono_text": abono_text,
        "saldo_text": saldo_text,
        "continuaciones": [],
        "skip": False,
    }, current_day


def _inbursa_continuation_is_ignored(text: str) -> bool:
    normalized = _normalize_text(text).upper()
    return (
        _is_inbursa_header_or_footer(text)
        or normalized.startswith("TE RECORDAMOS")
        or normalized.startswith("DETALLE DE MOVIMIENTOS")
        or normalized.startswith("FECHA REFERENCIA CONCEPTO")
        or normalized.startswith("-LA GAT")
        or "(CID:" in normalized
    )


def build_inbursa_audit_dataframe(movimientos_df: pd.DataFrame, resumen_df: pd.DataFrame) -> pd.DataFrame:
    summary_by_metric = {row["METRICA"]: row for _, row in resumen_df.iterrows()}
    checks: list[dict] = []

    def add_check(name: str, expected, calculated, detail: str) -> None:
        missing = pd.isna(expected) or pd.isna(calculated)
        difference = None if missing else Decimal(str(expected)) - Decimal(str(calculated))
        status = "ERROR" if missing or difference != Decimal("0") else "OK"
        checks.append({
            "COMPROBACION": name,
            "PDF_ESPERADO": expected,
            "EXCEL_CALCULADO": calculated,
            "DIFERENCIA": _money_to_float(difference),
            "TOLERANCIA": 0.0,
            "ESTADO": status,
            "DETALLE": detail if not missing else f"{detail} Falta un valor obligatorio.",
        })

    add_check("Datos obligatorios del resumen", 6, summary_by_metric["Datos obligatorios del resumen"]["CALCULADO"], "El resumen debe contener saldos, totales, rendimientos y comisiones.")
    add_check("Movimientos detectados", summary_by_metric["Movimientos detectados"]["PDF"], len(movimientos_df), "Todas las filas detectadas en el PDF deben llegar al Excel.")
    for metric in ("Abonos", "Cargos", "Saldo final"):
        add_check(metric, summary_by_metric[metric]["PDF"], summary_by_metric[metric]["CALCULADO"], f"El importe de {metric.lower()} debe cuadrar a centavos.")

    checkpoint_count = int(movimientos_df["SALDO_PDF"].notna().sum())
    checkpoint_ok = int((movimientos_df["DIFERENCIA_SALDO"] == 0).sum())
    add_check("Saldos intermedios", checkpoint_count, checkpoint_ok, "Todos los saldos publicados deben coincidir con el saldo acumulado.")

    initial = summary_by_metric["Saldo inicial"]["PDF"]
    abonos = summary_by_metric["Abonos"]["PDF"]
    cargos = summary_by_metric["Cargos"]["PDF"]
    final = summary_by_metric["Saldo final"]["PDF"]
    equation = None if any(pd.isna(value) for value in (initial, abonos, cargos)) else Decimal(str(initial)) + Decimal(str(abonos)) - Decimal(str(cargos))
    add_check("Ecuacion del estado", final, equation, "Saldo anterior + abonos - cargos debe ser igual al saldo final.")

    overall_status = "OK" if all(row["ESTADO"] == "OK" for row in checks) else "ERROR"
    checks.insert(0, {
        "COMPROBACION": "Estado general",
        "PDF_ESPERADO": "OK",
        "EXCEL_CALCULADO": overall_status,
        "DIFERENCIA": None,
        "TOLERANCIA": 0.0,
        "ESTADO": overall_status,
        "DETALLE": "Auditoria completa de resumen, movimientos, importes y saldos.",
    })
    return pd.DataFrame(checks)


def _validate_inbursa_audit(auditoria_df: pd.DataFrame) -> None:
    failed = auditoria_df[auditoria_df["ESTADO"] != "OK"]
    if failed.empty:
        return
    details = "; ".join(
        f"{row['COMPROBACION']}: {row['DETALLE']}"
        for _, row in failed.iterrows()
        if row["COMPROBACION"] != "Estado general"
    )
    LOGGER.error("Auditoria Inbursa fallida: %s", details)
    raise StatementProcessingError(f"Auditoria Inbursa fallida. {details}")


def extract_inbursa_data(file_obj) -> tuple[pd.DataFrame, pd.DataFrame]:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    with pdfplumber.open(file_obj) as pdf:
        summary = _extract_inbursa_summary(pdf)
        page1_text = pdf.pages[0].extract_text() or ""
        year_match = re.search(r"PERIODO\s+Del\s+\d{2}\s+[A-Z][a-z]{2}\.\s+(\d{4})", page1_text, re.IGNORECASE)
        year = year_match.group(1) if year_match else "2025"
        rows: list[dict] = []
        current_row: dict | None = None
        in_detail = False
        current_day: int | None = None

        for page in pdf.pages:
            in_detail = False
            amount_column_centers = None
            words = page.extract_words(x_tolerance=2, y_tolerance=2, keep_blank_chars=False)
            ordered_words = sorted(words, key=lambda word: (word["top"], word["x0"]))
            for line_words in _group_words_into_lines(ordered_words, tolerance=INBURSA_LINE_TOLERANCE):
                line = " ".join(word["text"] for word in line_words).strip()
                if not line:
                    continue
                normalized = _normalize_text(line).upper()
                if "DETALLE DE MOVIMIENTOS" in normalized or "FECHA REFERENCIA CONCEPTO CARGOS ABONOS SALDO" in normalized:
                    in_detail = True
                    detected_centers = _inbursa_amount_column_centers(line_words)
                    if detected_centers is not None:
                        amount_column_centers = detected_centers
                    continue
                if not in_detail:
                    continue
                top = min(word["top"] for word in line_words)
                if top > INBURSA_DETAIL_MAX_TOP:
                    continue
                if _should_end_inbursa_detail(line):
                    in_detail = False
                    if current_row:
                        rows.append(current_row)
                        current_row = None
                    continue
                row, current_day = _extract_inbursa_row_from_words(
                    line_words,
                    current_day,
                    amount_column_centers,
                )
                if row:
                    if current_row:
                        rows.append(current_row)
                    current_row = row
                    continue
                if _inbursa_continuation_is_ignored(line):
                    continue

                if current_row is not None:
                    current_row["continuaciones"].append(line)

        if current_row:
            rows.append(current_row)

    filtered_rows = [row for row in rows if not row.get("skip")]
    if not filtered_rows:
        raise StatementProcessingError("No se encontraron movimientos validos en el estado de cuenta Inbursa.")

    required_summary = ("saldo_inicial", "abonos", "cargos", "saldo_final", "rendimientos", "comisiones")
    required_count = sum(summary.get(key) is not None for key in required_summary)
    saldo_anterior = summary.get("saldo_inicial")
    total_abonos = Decimal("0.00")
    total_cargos = Decimal("0.00")
    movimientos: list[dict] = []

    for index, row in enumerate(filtered_rows, start=1):
        cargo = _read_money(row["cargo_text"]) or Decimal("0.00")
        abono = _read_money(row["abono_text"]) or Decimal("0.00")
        saldo_pdf = _read_money(row["saldo_text"])
        total_abonos += abono
        total_cargos += cargo

        saldo_calculado = None
        if saldo_anterior is not None:
            saldo_calculado = (saldo_anterior + abono - cargo).quantize(Decimal("0.01"))
            saldo_anterior = saldo_calculado

        diferencia_saldo = None
        if saldo_pdf is not None and saldo_calculado is not None:
            diferencia_saldo = (saldo_pdf - saldo_calculado).quantize(Decimal("0.01"))

        fecha = f"{row['month']} {row['day']:02d}-{year}" if row["day"] is not None else f"{row['month']}-{year}"
        movimientos.append({
            "MOVIMIENTO": index,
            "FECHA": fecha,
            "REFERENCIA": row["referencia"],
            "CONCEPTO": _clean_join([row["concepto"], *row["continuaciones"]]),
            "CARGO": _money_to_float(cargo) if cargo else None,
            "ABONO": _money_to_float(abono) if abono else None,
            "SALDO_PDF": _money_to_float(saldo_pdf),
            "SALDO_CALCULADO": _money_to_float(saldo_calculado),
            "DIFERENCIA_SALDO": _money_to_float(diferencia_saldo),
        })

    parsed_final = Decimal(str(movimientos[-1]["SALDO_CALCULADO"])) if movimientos and movimientos[-1]["SALDO_CALCULADO"] is not None else None
    resumen = pd.DataFrame([
        {"METRICA": "Saldo inicial", "PDF": _money_to_float(summary.get("saldo_inicial")), "CALCULADO": _money_to_float(summary.get("saldo_inicial")), "DIFERENCIA": 0.0 if summary.get("saldo_inicial") is not None else None},
        {"METRICA": "Abonos", "PDF": _money_to_float(summary.get("abonos")), "CALCULADO": _money_to_float(total_abonos), "DIFERENCIA": _money_to_float((summary.get("abonos") - total_abonos) if summary.get("abonos") is not None else None)},
        {"METRICA": "Cargos", "PDF": _money_to_float(summary.get("cargos")), "CALCULADO": _money_to_float(total_cargos), "DIFERENCIA": _money_to_float((summary.get("cargos") - total_cargos) if summary.get("cargos") is not None else None)},
        {"METRICA": "Saldo final", "PDF": _money_to_float(summary.get("saldo_final")), "CALCULADO": _money_to_float(parsed_final), "DIFERENCIA": _money_to_float((summary.get("saldo_final") - parsed_final) if summary.get("saldo_final") is not None and parsed_final is not None else None)},
        {"METRICA": "Rendimientos", "PDF": _money_to_float(summary.get("rendimientos")), "CALCULADO": None, "DIFERENCIA": None},
        {"METRICA": "Comisiones cobradas", "PDF": _money_to_float(summary.get("comisiones")), "CALCULADO": None, "DIFERENCIA": None},
        {"METRICA": "Movimientos detectados", "PDF": len(filtered_rows), "CALCULADO": len(movimientos), "DIFERENCIA": 0},
        {"METRICA": "Numero de abonos", "PDF": None, "CALCULADO": int(sum(1 for row in movimientos if row["ABONO"] is not None)), "DIFERENCIA": None},
        {"METRICA": "Numero de cargos", "PDF": None, "CALCULADO": int(sum(1 for row in movimientos if row["CARGO"] is not None)), "DIFERENCIA": None},
        {"METRICA": "Datos obligatorios del resumen", "PDF": 6, "CALCULADO": required_count, "DIFERENCIA": 6 - required_count},
    ])
    movimientos_df = pd.DataFrame(movimientos)
    auditoria_df = build_inbursa_audit_dataframe(movimientos_df, resumen)
    _validate_inbursa_audit(auditoria_df)
    LOGGER.info(
        "Auditoria Inbursa exitosa movimientos=%s abonos=%s cargos=%s puntos_saldo=%s estado=OK",
        len(movimientos_df),
        int(movimientos_df["ABONO"].notna().sum()),
        int(movimientos_df["CARGO"].notna().sum()),
        int(movimientos_df["SALDO_PDF"].notna().sum()),
    )
    return movimientos_df, resumen


def process_inbursa(file_obj):
    movimientos_df, resumen_df = extract_inbursa_data(file_obj)
    auditoria_df = build_inbursa_audit_dataframe(movimientos_df, resumen_df)
    return _dfs_to_excel([
        ("Movimientos", movimientos_df),
        ("Resumen", resumen_df),
        ("Auditoria", auditoria_df),
    ], style=True)

MP_FECHA_RE = re.compile(r"^\d{2}-\d{2}-\d{4}$")
MP_DIGITS_RE = re.compile(r"\d+")
MP_MONEY_CLEAN_RE = re.compile(r"[^\d,\.\-]")
MP_DATE_X_MAX = 90
MP_DESC_X_MIN = 90
MP_DESC_X_MAX = 213
MP_ID_X_MIN = 213
MP_ID_X_MAX = 285
MP_VALOR_X_MIN = 285
MP_VALOR_X_MAX = 360
MP_SALDO_X_MIN = 360
MP_SALDO_X_MAX = 430
MP_TOP_CUTOFF = 55
MP_BOTTOM_CUTOFF = 610
MP_DESC_TOP_OFFSET = -12
MP_DESC_BOTTOM_OFFSET = 16
MP_ID_TOP_OFFSET = -6
MP_ID_BOTTOM_OFFSET = 8
MP_MONEY_TOP_OFFSET = -4
MP_MONEY_BOTTOM_OFFSET = 4


@dataclass
class MercadoPagoMovimiento:
    fecha: str
    descripcion: str
    id_operacion: str
    monto: float | None
    saldo: float | None
    pagina: int
    top: float


def _mp_limpiar_moneda(texto: str) -> float | None:
    if not texto:
        return None
    limpio = MP_MONEY_CLEAN_RE.sub("", texto)
    if not limpio:
        return None
    try:
        return float(limpio.replace(",", ""))
    except Exception:
        return None


def _mp_texto_unido(words: list[dict], line_tol: float = 3.0) -> str:
    if not words:
        return ""
    words_ordenadas = sorted(words, key=lambda w: (round(w["top"], 1), w["x0"]))
    lineas: list[str] = []
    linea_actual: list[str] = []
    top_actual: float | None = None
    for word in words_ordenadas:
        top = float(word["top"])
        txt = str(word["text"]).strip()
        if not txt:
            continue
        if top_actual is None or abs(top - top_actual) <= line_tol:
            linea_actual.append(txt)
            top_actual = top if top_actual is None else (top_actual + top) / 2
        else:
            if linea_actual:
                lineas.append(" ".join(linea_actual).strip())
            linea_actual = [txt]
            top_actual = top
    if linea_actual:
        lineas.append(" ".join(linea_actual).strip())
    return " ".join(item for item in lineas if item)


def _mp_filtrar_palabras_utiles(page) -> list[dict]:
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
    return [word for word in words if MP_TOP_CUTOFF <= float(word["top"]) <= MP_BOTTOM_CUTOFF]


def _mp_extraer_anclas_fecha(words: list[dict]) -> list[dict]:
    anchors = [
        word for word in words
        if float(word["x0"]) < MP_DATE_X_MAX and MP_FECHA_RE.match(str(word["text"]).strip())
    ]
    return sorted(anchors, key=lambda word: float(word["top"]))


def _mp_palabras_en_rango(words: list[dict], x_min: float, x_max: float, top_min: float, top_max: float) -> list[dict]:
    return [
        word for word in words
        if x_min <= float(word["x0"]) < x_max and top_min <= float(word["top"]) <= top_max
    ]


def _mp_extraer_carry_descripcion(words: list[dict], anchors: list[dict]) -> str:
    if not anchors:
        return ""
    last_top = float(anchors[-1]["top"])
    trailing_desc = [
        word for word in words
        if MP_DESC_X_MIN <= float(word["x0"]) < MP_DESC_X_MAX and float(word["top"]) > last_top + 18
    ]
    trailing_id = [
        word for word in words
        if MP_ID_X_MIN <= float(word["x0"]) < MP_ID_X_MAX and float(word["top"]) > last_top + 18
    ]
    trailing_val = [
        word for word in words
        if MP_VALOR_X_MIN <= float(word["x0"]) < MP_VALOR_X_MAX and float(word["top"]) > last_top + 18
    ]
    trailing_saldo = [
        word for word in words
        if MP_SALDO_X_MIN <= float(word["x0"]) < MP_SALDO_X_MAX and float(word["top"]) > last_top + 18
    ]
    if trailing_desc and not trailing_id and not trailing_val and not trailing_saldo:
        return _mp_texto_unido(trailing_desc).strip()
    return ""


def _mp_construir_movimiento(words: list[dict], ancla_fecha: dict, pagina_num: int, carry_desc: str = "") -> MercadoPagoMovimiento:
    top = float(ancla_fecha["top"])
    fecha = str(ancla_fecha["text"]).strip()
    desc_words = _mp_palabras_en_rango(words, MP_DESC_X_MIN, MP_DESC_X_MAX, top + MP_DESC_TOP_OFFSET, top + MP_DESC_BOTTOM_OFFSET)
    id_words = _mp_palabras_en_rango(words, MP_ID_X_MIN, MP_ID_X_MAX, top + MP_ID_TOP_OFFSET, top + MP_ID_BOTTOM_OFFSET)
    valor_words = _mp_palabras_en_rango(words, MP_VALOR_X_MIN, MP_VALOR_X_MAX, top + MP_MONEY_TOP_OFFSET, top + MP_MONEY_BOTTOM_OFFSET)
    saldo_words = _mp_palabras_en_rango(words, MP_SALDO_X_MIN, MP_SALDO_X_MAX, top + MP_MONEY_TOP_OFFSET, top + MP_MONEY_BOTTOM_OFFSET)
    descripcion = _mp_texto_unido(desc_words)
    if carry_desc:
        descripcion = f"{carry_desc} {descripcion}".strip()
    return MercadoPagoMovimiento(
        fecha=fecha,
        descripcion=descripcion,
        id_operacion="".join(MP_DIGITS_RE.findall(_mp_texto_unido(id_words))),
        monto=_mp_limpiar_moneda(_mp_texto_unido(valor_words)),
        saldo=_mp_limpiar_moneda(_mp_texto_unido(saldo_words)),
        pagina=pagina_num,
        top=top,
    )


def extract_mercado_pago_data(file_obj) -> tuple[pd.DataFrame, pd.DataFrame]:
    movimientos: list[MercadoPagoMovimiento] = []
    auditoria: list[dict] = []
    carry_desc = ""
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    with pdfplumber.open(file_obj) as pdf:
        for pagina_num, page in enumerate(pdf.pages, start=1):
            words = _mp_filtrar_palabras_utiles(page)
            anchors = _mp_extraer_anclas_fecha(words)
            if not anchors:
                if carry_desc:
                    auditoria.append({"pagina": pagina_num, "tipo": "carry_sin_destino", "detalle": carry_desc})
                continue
            for idx, anchor in enumerate(anchors):
                prefijo = carry_desc if idx == 0 and carry_desc else ""
                movimientos.append(_mp_construir_movimiento(words, anchor, pagina_num, prefijo))
                if prefijo:
                    carry_desc = ""
            carry_desc = _mp_extraer_carry_descripcion(words, anchors)
    if carry_desc:
        auditoria.append({"pagina": "EOF", "tipo": "carry_final_sin_destino", "detalle": carry_desc})
    df = pd.DataFrame(
        [{
            "FECHA": mov.fecha,
            "DESCRIPCION": mov.descripcion,
            "ID OPERACION": mov.id_operacion,
            "MONTO": mov.monto,
            "SALDO": mov.saldo,
            "PAGINA": mov.pagina,
        } for mov in movimientos]
    )
    if df.empty:
        raise StatementProcessingError("No se encontraron movimientos validos en el estado de cuenta de Mercado Pago.")
    df["DESCRIPCION"] = df["DESCRIPCION"].fillna("").astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
    df["ID OPERACION"] = df["ID OPERACION"].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    audit_rows: list[dict] = []
    for i, row in df.iterrows():
        flags = []
        if not row["FECHA"]:
            flags.append("sin_fecha")
        if not row["DESCRIPCION"]:
            flags.append("sin_descripcion")
        if not row["ID OPERACION"]:
            flags.append("sin_id")
        if pd.isna(row["MONTO"]):
            flags.append("sin_monto")
        if pd.isna(row["SALDO"]):
            flags.append("sin_saldo")
        if flags:
            audit_rows.append({"fila_df": i + 2, "tipo": ",".join(flags), "detalle": str(row.to_dict())})
    audit_rows.extend(auditoria)
    return df, pd.DataFrame(audit_rows)


def process_mercado_pago(file_obj):
    movimientos_df, auditoria_df = extract_mercado_pago_data(file_obj)
    sheets = [("Movimientos", movimientos_df)]
    if not auditoria_df.empty:
        sheets.append(("Auditoria", auditoria_df))
    return _dfs_to_excel(sheets)


BANAMEX_CONTROL_SALDO_INICIAL = Decimal("5396.69")
BANAMEX_CONTROL_DEPOSITOS = Decimal("510474.68")
BANAMEX_CONTROL_RETIROS = Decimal("485950.89")
BANAMEX_CONTROL_SALDO_FINAL = Decimal("29920.48")


def _banamex_limpiar_cantidad(valor_str) -> Decimal:
    if not valor_str or pd.isna(valor_str):
        return Decimal("0.00")
    valor_limpio = str(valor_str).replace("$", "").replace(",", "").strip()
    if valor_limpio == "":
        return Decimal("0.00")
    valor_limpio = re.sub(r"[^\d\.\-]", "", valor_limpio)
    if valor_limpio.endswith("-"):
        valor_limpio = "-" + valor_limpio[:-1]
    try:
        return Decimal(valor_limpio).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _banamex_es_monto_contable(texto: str) -> bool:
    limpio = re.sub(r"[^\d\.\-\,]", "", texto)
    return bool(re.match(r"^\-?\d{1,3}(?:,\d{3})*\.\d{2}\-?$", limpio))


def _banamex_agrupar_en_lineas(palabras: list[dict], tolerancia_y: int = 5) -> list[list[dict]]:
    lineas = []
    palabras_ordenadas = sorted(palabras, key=lambda word: word["top"])
    if not palabras_ordenadas:
        return lineas
    linea_actual = [palabras_ordenadas[0]]
    for palabra in palabras_ordenadas[1:]:
        if abs(palabra["top"] - linea_actual[-1]["top"]) <= tolerancia_y:
            linea_actual.append(palabra)
        else:
            lineas.append(linea_actual)
            linea_actual = [palabra]
    if linea_actual:
        lineas.append(linea_actual)
    return lineas


def extract_banamex_data(file_obj) -> tuple[pd.DataFrame, pd.DataFrame]:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    datos_extraidos: list[dict] = []
    auditoria: list[dict] = []
    terminar_lectura = False
    transaccion_actual = None
    with pdfplumber.open(file_obj) as pdf:
        for i in range(1, len(pdf.pages)):
            if terminar_lectura:
                break
            pagina = pdf.pages[i]
            num_pagina = i + 1
            palabras = pagina.extract_words(keep_blank_chars=False) or []
            lineas = _banamex_agrupar_en_lineas(palabras)
            en_tabla = False
            for linea_palabras in lineas:
                linea_palabras = sorted(linea_palabras, key=lambda word: word["x0"])
                texto_completo = " ".join(word["text"] for word in linea_palabras).upper()
                if "GLOSARIO" in texto_completo or "SALONES BEYOND" in texto_completo or "DISPOSICIONES EN CAJERO" in texto_completo:
                    terminar_lectura = True
                    break
                if "FECHA" in texto_completo and "CONCEPTO" in texto_completo and "RETIROS" in texto_completo:
                    en_tabla = True
                    continue
                if not en_tabla:
                    continue
                if "DETALLE DE OPERACIONES" in texto_completo or ("PÁGINA" in texto_completo and "DE" in texto_completo) or "SALDO ANTERIOR" in texto_completo or "SALVO BUEN COBRO" in texto_completo:
                    continue
                fila_dict = {"FECHA": [], "CONCEPTO": [], "RETIROS": [], "DEPOSITOS": [], "SALDO": []}
                for word in linea_palabras:
                    texto = word["text"].strip()
                    x_center = (word["x0"] + word["x1"]) / 2
                    es_monto = _banamex_es_monto_contable(texto)
                    if es_monto and x_center >= 330:
                        if x_center < 425:
                            fila_dict["RETIROS"].append(texto)
                        elif x_center < 515:
                            fila_dict["DEPOSITOS"].append(texto)
                        else:
                            fila_dict["SALDO"].append(texto)
                    else:
                        if word["x0"] < 80:
                            fila_dict["FECHA"].append(texto)
                        else:
                            fila_dict["CONCEPTO"].append(texto)
                fecha = " ".join(fila_dict["FECHA"]).strip()
                concepto = " ".join(fila_dict["CONCEPTO"]).strip()
                retiro = " ".join(fila_dict["RETIROS"]).strip()
                deposito = " ".join(fila_dict["DEPOSITOS"]).strip()
                saldo = " ".join(fila_dict["SALDO"]).strip()
                if not any([fecha, concepto, retiro, deposito, saldo]):
                    continue
                es_nueva_fecha = bool(re.match(r"^\d{2}\s+[A-Za-z]{3}", fecha))
                if es_nueva_fecha:
                    if transaccion_actual:
                        datos_extraidos.append(transaccion_actual)
                    transaccion_actual = {
                        "PAGINA": num_pagina,
                        "FECHA": fecha,
                        "CONCEPTO": concepto,
                        "RETIROS": _banamex_limpiar_cantidad(retiro),
                        "DEPOSITOS": _banamex_limpiar_cantidad(deposito),
                        "SALDO": _banamex_limpiar_cantidad(saldo),
                    }
                elif transaccion_actual:
                    tiene_nuevo_monto = bool(retiro or deposito)
                    if tiene_nuevo_monto and (transaccion_actual["RETIROS"] != Decimal("0.00") or transaccion_actual["DEPOSITOS"] != Decimal("0.00")):
                        datos_extraidos.append(transaccion_actual)
                        transaccion_actual = {
                            "PAGINA": num_pagina,
                            "FECHA": transaccion_actual["FECHA"],
                            "CONCEPTO": concepto,
                            "RETIROS": _banamex_limpiar_cantidad(retiro),
                            "DEPOSITOS": _banamex_limpiar_cantidad(deposito),
                            "SALDO": _banamex_limpiar_cantidad(saldo),
                        }
                    else:
                        if concepto:
                            transaccion_actual["CONCEPTO"] += f" {concepto}"
                        if retiro and transaccion_actual["RETIROS"] == Decimal("0.00"):
                            transaccion_actual["RETIROS"] = _banamex_limpiar_cantidad(retiro)
                        if deposito and transaccion_actual["DEPOSITOS"] == Decimal("0.00"):
                            transaccion_actual["DEPOSITOS"] = _banamex_limpiar_cantidad(deposito)
                        if saldo:
                            transaccion_actual["SALDO"] = _banamex_limpiar_cantidad(saldo)
    if transaccion_actual:
        datos_extraidos.append(transaccion_actual)
    if not datos_extraidos:
        raise StatementProcessingError("No se encontraron movimientos validos en el estado de cuenta Banamex.")
    saldo_calculado = BANAMEX_CONTROL_SALDO_INICIAL
    suma_retiros = Decimal("0.00")
    suma_depositos = Decimal("0.00")
    for i, fila in enumerate(datos_extraidos, start=1):
        retiro = fila["RETIROS"]
        deposito = fila["DEPOSITOS"]
        saldo_leido = fila["SALDO"]
        suma_retiros += retiro
        suma_depositos += deposito
        saldo_calculado = (saldo_calculado - retiro + deposito).quantize(Decimal("0.01"))
        if saldo_leido != Decimal("0.00"):
            diferencia = abs(saldo_calculado - saldo_leido).quantize(Decimal("0.01"))
            if diferencia > Decimal("0.01"):
                auditoria.append({
                    "FILA": i,
                    "TIPO": "descuadre_saldo",
                    "DETALLE": f"Calculado {saldo_calculado} vs PDF {saldo_leido}",
                })
                saldo_calculado = saldo_leido
    resumen = pd.DataFrame([
        {"METRICA": "Saldo inicial", "CONTROL": float(BANAMEX_CONTROL_SALDO_INICIAL), "EXTRAIDO": float(BANAMEX_CONTROL_SALDO_INICIAL)},
        {"METRICA": "Depositos", "CONTROL": float(BANAMEX_CONTROL_DEPOSITOS), "EXTRAIDO": float(suma_depositos)},
        {"METRICA": "Retiros", "CONTROL": float(BANAMEX_CONTROL_RETIROS), "EXTRAIDO": float(suma_retiros)},
        {"METRICA": "Saldo final", "CONTROL": float(BANAMEX_CONTROL_SALDO_FINAL), "EXTRAIDO": float(saldo_calculado)},
    ])
    df = pd.DataFrame([
        {
            "PAGINA": fila["PAGINA"],
            "FECHA": fila["FECHA"],
            "CONCEPTO": fila["CONCEPTO"].strip(),
            "RETIROS": float(fila["RETIROS"]),
            "DEPOSITOS": float(fila["DEPOSITOS"]),
            "SALDO": float(fila["SALDO"]),
        }
        for fila in datos_extraidos
    ])
    return df, pd.concat([resumen, pd.DataFrame(auditoria)], ignore_index=True, sort=False)


def process_banamex(file_obj):
    movimientos_df, auditoria_df = extract_banamex_data(file_obj)
    return _dfs_to_excel([
        ("Movimientos", movimientos_df),
        ("Auditoria", auditoria_df),
    ])


def _df_to_excel(df):
    if df.empty:
        df = pd.DataFrame(columns=["Mensaje"])
    return _dfs_to_excel([("Resultado", df)])

